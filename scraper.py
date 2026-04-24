#!/usr/bin/env python3
"""
Google Maps Scraper - Core Scraping Logic
Robust scraper using Playwright for extracting business information
"""

import asyncio
import re
import urllib.parse
import random
import os
import json
from typing import List, Dict, Optional
import pandas as pd
from playwright.async_api import async_playwright, Browser, Page, BrowserContext
from langchain_groq import ChatGroq
from langchain_core.messages import SystemMessage, HumanMessage


class GoogleMapsScraper:
    """Google Maps scraper with anti-detection measures"""

    def __init__(
        self,
        headless: bool = False,
        delay_range: tuple = (1, 3),
        max_results: int = 50,
        timeout: int = 30000
    ):
        self.headless = headless
        self.delay_range = delay_range
        self.max_results = max_results
        self.timeout = timeout
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.page: Optional[Page] = None
        
        # Custom user agent to mimic real Chrome browser
        self.user_agent = (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )

    async def init_browser(self):
        """Initialize browser with anti-detection settings"""
        playwright = await async_playwright().start()
        
        self.browser = await playwright.chromium.launch(
            headless=self.headless,
            channel="chrome"
        )
        
        self.context = await self.browser.new_context(
            user_agent=self.user_agent,
            viewport={"width": 1920, "height": 1080},
            locale="en-US",
            timezone_id="America/New_York"
        )
        
        # Add stealth scripts to avoid detection
        await self.context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            
            window.chrome = {
                runtime: {},
                loadTimes: function() {
                    return {
                        commitLoadTime: performance.timing.domContentLoadedEventStart / 1000,
                        connectionInfo: 'h2',
                        finishDocumentLoadTime: performance.timing.domContentLoadedEventEnd / 1000,
                        finishLoadTime: performance.timing.loadEventEnd / 1000,
                        firstPaintAfterLoadTime: 0,
                        firstPaintTime: 0,
                        navigationType: 'Other',
                        npnNegotiatedProtocol: 'h2',
                        requestTime: performance.timing.requestStart / 1000,
                        startLoadTime: performance.timing.requestStart / 1000,
                        wasAlternateProtocolAvailable: false,
                        wasFetchedViaSpdy: true,
                        wasNpnNegotiated: true
                    };
                },
                csi: function() {
                    return {
                        onloadT: Date.now(),
                        pageT: Date.now() - performance.timing.navigationStart,
                        startE: performance.timing.navigationStart,
                        tran: 15
                    };
                }
            };
            
            delete navigator.__proto__.webdriver;
        """)
        
        self.page = await self.context.new_page()
        
        # Set extra HTTP headers
        await self.page.set_extra_http_headers({
            "Accept-Language": "en-US,en;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "DNT": "1",
            "Connection": "keep-alive",
            "Upgrade-Insecure-Requests": "1"
        })

    async def random_delay(self):
        """Add random delay between interactions"""
        delay = random.uniform(*self.delay_range)
        await asyncio.sleep(delay)

    async def scroll_to_load_all_results(self):
        """Auto-scroll the results sidebar to load all available businesses"""
        if not self.page:
            return
            
        print("📜 Scrolling to load all results...")

        try:
            feed = await self.page.wait_for_selector('div[role="feed"]', timeout=10000)

            previous_count = 0
            same_count_iterations = 0
            max_same_count = 6

            while same_count_iterations < max_same_count:
                await feed.evaluate("el => { el.scrollTop = el.scrollHeight; }")
                await self.random_delay()

                current_count = await self.page.locator('div[role="article"], [data-result-index]').count()

                if current_count > previous_count:
                    previous_count = current_count
                    same_count_iterations = 0
                    print(f"📊 Loaded {current_count} results so far...")
                else:
                    same_count_iterations += 1

                if current_count >= self.max_results:
                    break
                        
        except Exception as e:
            print(f"⚠️  Error during scrolling: {e}")

    async def extract_business_info(self, business_element) -> Dict:
        """Extract business information from a single business element"""
        business_data = {
            "name": "",
            "rating": "",
            "reviews_count": "",
            "location_category": "",
            "facility_category": "",
            "address": "",
            "website": "",
            "email": "",
            "phone": "",
            "number_division": "",
            "PIC": "Scrapper by AI"
        }

        try:
            # Extract name
            name_selectors = [
                '.fontHeadlineSmall',
                'h3',
                '[data-result-index] h3',
                '.section-result-title span',
                'div[role="heading"]'
            ]
            
            for selector in name_selectors:
                try:
                    name_element = await business_element.query_selector(selector)
                    if name_element:
                        text = await name_element.text_content()
                        if text and text.strip():
                            business_data["name"] = text.strip()
                            break
                except:
                    continue

            # If name is still empty, try to get the aria-label of the main element
            if not business_data["name"]:
                try:
                    aria_label = await business_element.get_attribute("aria-label")
                    if aria_label:
                        business_data["name"] = aria_label.strip()
                except:
                    pass

            # Only proceed to extract other details if we have a valid name
            # and the name doesn't look like a whole block of text
            if business_data["name"] and len(business_data["name"]) < 100:
                
                # Extract rating
                rating_selectors = [
                    'span[role="img"]',
                    '.section-result-rating',
                    '.rating-text',
                    '[data-result-index] span[role="img"]'
                ]
                
                for selector in rating_selectors:
                    try:
                        rating_element = await business_element.query_selector(selector)
                        if rating_element:
                            rating_text = await rating_element.get_attribute("aria-label") or await rating_element.text_content()
                            if rating_text and ("star" in rating_text.lower() or "bintang" in rating_text.lower() or "." in rating_text or "," in rating_text):
                                rating_text = rating_text.replace(',', '.')
                                rating_match = re.search(r'(\d+\.\d+)', rating_text)
                                if rating_match:
                                    business_data["rating"] = rating_match.group(1)
                                    review_match = re.search(r'(?i)(\d+(?:\.\d+)?)\s*(?:Ulasan|Reviews)', rating_text.replace('.', ''))
                                    if review_match and not business_data["reviews_count"]:
                                        business_data["reviews_count"] = review_match.group(1)
                                    break
                    except:
                        continue

                # Extract reviews count
                reviews_selectors = [
                    'span[role="img"] + span',
                    '.section-result-rating + span',
                    '.reviews-text',
                    'span:has-text("(")'
                ]
                
                for selector in reviews_selectors:
                    try:
                        reviews_element = await business_element.query_selector(selector)
                        if reviews_element:
                            reviews_text = await reviews_element.text_content()
                            if reviews_text and "(" in reviews_text:
                                clean_text = reviews_text.replace('.', '').replace(',', '')
                                reviews_match = re.search(r'\((\d+)\)', clean_text)
                                if reviews_match:
                                    business_data["reviews_count"] = reviews_match.group(1)
                                    break
                    except:
                        continue

                # Extract category and address from the text block
                try:
                    # Look for specific element classes that hold category/address
                    font_elements = await business_element.query_selector_all('.fontBodyMedium')
                    
                    for element in font_elements:
                        text = await element.text_content()
                        if text and "·" in text:
                            # Clean up the text first to remove strange characters
                            clean_text = re.sub(r'[\ue000-\uf8ff]', '', text).strip()
                            parts = [p.strip() for p in clean_text.split("·")]
                            
                            # Sometimes the reviews are hiding in the first part like "4,6(58)"
                            if not business_data["reviews_count"]:
                                for part in parts:
                                    clean_part = part.replace('.', '').replace(',', '')
                                    rev_match = re.search(r'\((\d+)\)', clean_part)
                                    if rev_match:
                                        business_data["reviews_count"] = rev_match.group(1)
                                        break
                                        
                            # Usually the structure is: [Rating (Reviews)] · [Category] · [Address/Location]
                            # Example: "5.0 (12) · Kedai Kopi · Jl. ABC"
                            
                            for i, part in enumerate(parts):
                                # Skip parts that are just ratings/reviews like "5,0(12)"
                                if re.match(r'^[\d\.,\s]+(?:\([\d\.,\s]+\))?$', part):
                                    continue
                                
                                # First text-heavy part is likely the category
                                if not business_data["location_category"] and len(part) < 50:
                                    # Ensure it's not grabbing the business name or rating
                                    if part != business_data["name"] and not re.match(r'^[\d\.,\s]+(?:\([\d\.,\s]+\))?$', part):
                                        # Sometimes it merges name + rating + category like "First Crack Coffee Rajawali Place  5,0Kedai Kopi"
                                        # Let's clean it up by only taking the actual category words
                                        clean_cat = re.sub(r'^.*?\s+[\d\.,]+\s*', '', part)
                                        
                                        # If the category looks like an address, it means category was missing entirely
                                        if re.search(r'(?i)(Jl\.|Jalan|Kav|RT|RW)', clean_cat):
                                            business_data["address"] = clean_cat.strip()
                                        elif not re.search(r'(?i)(Buka|Tutup|pukul)', clean_cat):
                                            # Avoid things like "Buka Kam pukul 08.30" as category
                                            business_data["location_category"] = clean_cat.strip() if clean_cat else part.strip()
                                    continue
                                
                                # Next part is likely the address
                                if business_data["location_category"] and not business_data["address"]:
                                    # Clean up common Google Maps text like "Buka", "Tutup", "Pesan", "Segera tutup"
                                    clean_address = re.sub(r'(Buka|Tutup|Pesan|Segera\s*tutup|Segera).*$', '', part, flags=re.IGNORECASE).strip()
                                    if clean_address and len(clean_address) > 5:
                                        business_data["address"] = clean_address
                                        break
                                        
                            if business_data["location_category"] and business_data["address"]:
                                break
                except:
                    pass
                
                # Fallback for category and address if the previous method failed
                if not business_data["location_category"] or not business_data["address"]:
                    try:
                        # Extract the aria-label of the entire element, it often contains the structured info
                        aria_label = await business_element.get_attribute("aria-label")
                        if aria_label:
                            # Typical format: "Gould Coffee & Eatery Setiabudi 4,6(58) · Kedai Kopi ·  · Buka ⋅ Tutup pukul 21.00"
                            
                            # Clean the text
                            clean_label = re.sub(r'[\ue000-\uf8ff]', '', aria_label).strip()
                            parts = [p.strip() for p in clean_label.split("·")]
                            
                            for part in parts:
                                # If it's the name or a rating, skip
                                if business_data["name"] in part or re.match(r'^[\d\.,\s]+(?:\([\d\.,\s]+\))?$', part):
                                    continue
                                
                                # Category fallback
                                if not business_data["location_category"] and len(part) < 40 and not "Buka" in part and not "Tutup" in part:
                                    business_data["location_category"] = part
                                    continue
                                    
                                # Address fallback
                                if not business_data["address"] and len(part) > 5:
                                    clean_address = re.sub(r'(Buka|Tutup|Pesan|Segera).*$', '', part, flags=re.IGNORECASE).strip()
                                    if clean_address and clean_address != business_data["location_category"]:
                                        business_data["address"] = clean_address
                                        break
                    except:
                        pass

        except Exception as e:
            print(f"⚠️  Error extracting business info: {e}")
            
        # Final cleanup of the extracted strings
        for key, value in business_data.items():
            if value and isinstance(value, str):
                # Remove common Google Maps garbage characters and normalize spaces
                clean_val = re.sub(r'[\ue000-\uf8ff\u200e]', '', value).strip()
                # Remove "Buka" or "Tutup" from addresses if it slipped through
                if key == "address":
                    clean_val = re.sub(r'(?i)\s*(Buka|Tutup|Pesan|Segera).*$', '', clean_val).strip()
                business_data[key] = clean_val

        return business_data

    def categorize_facilities_with_ai(self, results: List[Dict]) -> List[Dict]:
        return self.categorize_location_and_facility_with_ai(results)

    def _infer_facility_category_from_location(self, location_category: str) -> str:
        t = (location_category or "").strip().lower()
        if not t:
            return "Fasilitas Umum"
        if any(k in t for k in ["apartemen", "apartment", "residence", "residences", "condo", "kondominium"]):
            return "Apartemen"
        if any(k in t for k in ["rumah sakit", "hospital", "klinik", "clinic", "puskesmas", "medical", "health", "kesehatan"]):
            return "Sarana Kesehatan"
        if any(k in t for k in ["gym", "fitness", "stadion", "stadium", "lapangan", "sport", "olahraga", "basket", "futsal", "badminton", "tenis", "swimming", "renang"]):
            return "Sarana Olahraga"
        return "Fasilitas Umum"

    def _is_valid_location_category(self, value: str) -> bool:
        v = (value or "").strip()
        if not v:
            return False
        if len(v) > 40:
            return False
        if any(ch.isdigit() for ch in v):
            return False
        lower = v.lower()
        forbidden = [
            "http", "@", "buka", "tutup", "pukul", "open", "close", "hours",
            "jl", "jalan", "rt", "rw", "kav", "no.", "blok", "block", "kecamatan", "kelurahan"
        ]
        if any(k in lower for k in forbidden):
            return False
        if any(k in v for k in [",", "·", "|"]):
            return False
        words = [w for w in re.split(r"\s+", v) if w]
        if len(words) > 5:
            return False
        return True

    def categorize_location_and_facility_with_ai(self, results: List[Dict]) -> List[Dict]:
        if not results:
            return results

        if not os.getenv("GROQ_API_KEY"):
            return results

        llm = ChatGroq(
            model="meta-llama/llama-4-scout-17b-16e-instruct",
            temperature=0
        )

        allowed_facility = ["Sarana Olahraga", "Sarana Kesehatan", "Apartemen", "Fasilitas Umum"]

        batch_size = 25
        print("🤖 AI categorization: LOCATION_CATEGORY + FACILITY_CATEGORY")
        for start in range(0, len(results), batch_size):
            batch = results[start:start + batch_size]
            payload = []
            for idx, item in enumerate(batch, start=1):
                payload.append({
                    "idx": idx,
                    "name": item.get("name", ""),
                    "location_category": item.get("location_category", ""),
                    "address": item.get("address", ""),
                    "website": item.get("website", "")
                })

            system = (
                "Kamu adalah classifier untuk data tempat dari Google Maps. "
                "Tugasmu untuk tiap item:\n"
                "1) Isi location_category (jenis tempat) yang singkat dan rapi (1-4 kata), contoh: Universitas, Kedai Kopi, Apartemen, Rumah Sakit.\n"
                "2) Isi facility_category yang harus salah satu dari: "
                + ", ".join(allowed_facility)
                + ". Jika ragu, pilih Fasilitas Umum.\n"
                "Jangan masukkan rating, jam buka, atau alamat ke location_category."
            )

            human = (
                "Kembalikan JSON array saja (tanpa teks tambahan). "
                "Format: [{\"idx\":1,\"location_category\":\"Universitas\",\"facility_category\":\"Fasilitas Umum\"}, ...]. "
                "Berikut datanya:\n"
                + json.dumps(payload, ensure_ascii=False)
            )

            try:
                response = llm.invoke([SystemMessage(content=system), HumanMessage(content=human)])
                text = (response.content or "").strip()

                try:
                    parsed = json.loads(text)
                except Exception:
                    left = text.find("[")
                    right = text.rfind("]")
                    parsed = json.loads(text[left:right + 1]) if left != -1 and right != -1 and right > left else []

                mapping = {}
                if isinstance(parsed, list):
                    for row in parsed:
                        if not isinstance(row, dict) or "idx" not in row:
                            continue
                        idx = int(row["idx"])
                        mapping[idx] = {
                            "location_category": str(row.get("location_category", "") or "").strip(),
                            "facility_category": str(row.get("facility_category", "") or "").strip()
                        }

                for idx, item in enumerate(batch, start=1):
                    picked = mapping.get(idx, {})
                    loc = (picked.get("location_category") or "").strip()
                    fac = (picked.get("facility_category") or "").strip()

                    if loc:
                        loc = re.sub(r'(?i)\s*(Buka|Tutup|Pesan|Segera).*$', '', loc).strip()
                        if self._is_valid_location_category(loc):
                            item["location_category"] = loc

                    if fac not in allowed_facility:
                        fac = self._infer_facility_category_from_location(item.get("location_category", ""))
                    item["facility_category"] = fac
            except Exception:
                for item in batch:
                    if not item.get("facility_category"):
                        item["facility_category"] = self._infer_facility_category_from_location(item.get("location_category", ""))

        return results

    async def _search_website_for_email(self, website_url: str) -> str:
        if not website_url or not self.context:
            return ""

        page = await self.context.new_page()
        try:
            await page.goto(website_url, wait_until="domcontentloaded", timeout=20000)
            await asyncio.sleep(1)

            mailto = await page.query_selector('a[href^="mailto:"]')
            if mailto:
                href = await mailto.get_attribute("href")
                if href and href.lower().startswith("mailto:"):
                    email = href.split(":", 1)[1].split("?", 1)[0].strip()
                    if email and re.match(r'^[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}$', email, re.IGNORECASE):
                        return email

            body_text = await page.evaluate("document.body ? document.body.innerText : ''")
            if body_text:
                match = re.search(r'[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}', body_text, re.IGNORECASE)
                if match:
                    return match.group(0).strip()

            link = await page.query_selector('a[href*="contact"], a[href*="kontak"], a:has-text("Contact"), a:has-text("Kontak")')
            if link:
                href = await link.get_attribute("href")
                if href:
                    try:
                        await page.goto(href, wait_until="domcontentloaded", timeout=20000)
                        await asyncio.sleep(1)
                        body_text = await page.evaluate("document.body ? document.body.innerText : ''")
                        if body_text:
                            match = re.search(r'[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}', body_text, re.IGNORECASE)
                            if match:
                                return match.group(0).strip()
                    except Exception:
                        pass
        except Exception:
            return ""
        finally:
            try:
                await page.close()
            except Exception:
                pass

        return ""

    def _extract_email(self, text: str) -> str:
        if not text:
            return ""
        match = re.search(r'[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}', text, re.IGNORECASE)
        return match.group(0).strip() if match else ""

    def _infer_number_division(self, text: str) -> str:
        if not text:
            return ""
        t = re.sub(r"\s+", " ", text).strip().lower()
        candidates = [
            ("Marketing", ["marketing", "pemasaran", "promo", "promotion"]),
            ("Business Development", ["business development", "bizdev", "b.d", "partnership", "partnerships", "kerjasama", "kerja sama", "kemitraan", "cooperation"]),
            ("Humas", ["humas", "public relations", "corporate communication", "corporate communications", "communication", "komunikasi", "corcomm", "media", "press", "pr"]),
        ]
        best = ""
        best_score = 0
        for label, keys in candidates:
            score = 0
            for k in keys:
                if k in t:
                    score += 1
            if score > best_score:
                best_score = score
                best = label
        return best if best_score > 0 else ""

    def _pick_preferred_contacts_from_text(self, text: str) -> Dict[str, str]:
        if not text:
            return {"phone": "", "email": "", "phone_division": "", "email_division": ""}

        normalized = re.sub(r"\s+", " ", text)
        lower = normalized.lower()

        def div_priority(div: str) -> int:
            if div == "Marketing":
                return 3
            if div == "Business Development":
                return 2
            if div == "Humas":
                return 1
            return 0

        def division_for_span(start: int, end: int, value: str) -> tuple[str, int]:
            left = max(0, start - 90)
            right = min(len(lower), end + 90)
            context = lower[left:right]
            div = self._infer_number_division(context)
            score = 0
            if div:
                score += 3
            div2 = self._infer_number_division(value)
            if div2 and div2 != div:
                div = div2
                score += 3
            elif div2:
                div = div2
                score += 2
            return div, score

        best_email = ""
        best_email_div = ""
        best_email_score = -1

        for m in re.finditer(r'[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}', normalized, re.IGNORECASE):
            email = m.group(0).strip()
            div, score = division_for_span(m.start(), m.end(), email)
            if score > best_email_score or (score == best_email_score and div_priority(div) > div_priority(best_email_div)):
                best_email = email
                best_email_div = div
                best_email_score = score

        best_phone = ""
        best_phone_div = ""
        best_phone_score = -1

        for m in re.finditer(r'(?:\+62\s*8|0\s*8)[\d\s\-\.]{8,18}', normalized):
            raw = m.group(0)
            phone = re.sub(r'[^\d\+]', '', raw)
            if not (phone.startswith("+628") or phone.startswith("08")):
                continue
            div, score = division_for_span(m.start(), m.end(), raw)
            if score > best_phone_score or (score == best_phone_score and div_priority(div) > div_priority(best_phone_div)):
                best_phone = phone
                best_phone_div = div
                best_phone_score = score

        return {"phone": best_phone, "email": best_email, "phone_division": best_phone_div, "email_division": best_email_div}

    def _extract_mobile_phone(self, text: str) -> str:
        if not text:
            return ""
        match = re.search(r'(?:\+62\s*8|0\s*8)[\d\s\-\.]{8,18}', text)
        if not match:
            return ""
        raw = match.group(0)
        num = re.sub(r'[^\d\+]', '', raw)
        if num.startswith("0"):
            num = num
        if num.startswith("+628") or num.startswith("08"):
            return num
        return ""

    async def _search_web_for_contacts(self, business_name: str, website_url: str = "") -> Dict[str, str]:
        if not self.context:
            return {"phone": "", "email": "", "phone_division": "", "email_division": ""}

        print(f"    🌐 Searching web for contacts: {business_name}...")

        page = await self.context.new_page()
        phone = ""
        email = ""
        phone_division = ""
        email_division = ""
        
        try:
            search_query = urllib.parse.quote(
                f"{business_name} (marketing OR humas OR \"business development\" OR kerjasama) (kontak OR contact) (nomor OR whatsapp OR email)"
            )
            search_url = f"https://www.google.com/search?q={search_query}"
            
            await page.goto(search_url, wait_until="domcontentloaded", timeout=20000)

            snippet_elements = await page.query_selector_all('.VwiC3b, .yXK7lf, .MUxGbd, .aCOpRe')
            for snippet in snippet_elements:
                text = await snippet.text_content()
                if text:
                    picked = self._pick_preferred_contacts_from_text(text)
                    if picked.get("email") and (not email or (not email_division and picked.get("email_division"))):
                        email = picked["email"]
                        email_division = picked.get("email_division", "")
                    if picked.get("phone") and (not phone or (not phone_division and picked.get("phone_division"))):
                        phone = picked["phone"]
                        phone_division = picked.get("phone_division", "")
                    if phone and email:
                        break

            if not email:
                mailto = await page.query_selector('a[href^="mailto:"]')
                if mailto:
                    href = await mailto.get_attribute("href")
                    if href and href.lower().startswith("mailto:"):
                        candidate = href.split(":", 1)[1].split("?", 1)[0].strip()
                        if candidate and re.match(r'^[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}$', candidate, re.IGNORECASE):
                            email = candidate
                            email_division = self._infer_number_division(candidate)

            if not phone or not email:
                result_links = await page.query_selector_all('a[href^="http"]')
                visited = set()

                if website_url and isinstance(website_url, str) and website_url.startswith("http"):
                    try:
                        await page.goto(website_url, wait_until="domcontentloaded", timeout=20000)
                        await asyncio.sleep(1)
                        body_text = await page.evaluate("document.body ? document.body.innerText : ''")
                        picked = self._pick_preferred_contacts_from_text(body_text or "")
                        if picked.get("email") and (not email or (not email_division and picked.get("email_division"))):
                            email = picked["email"]
                            email_division = picked.get("email_division", "")
                        if picked.get("phone") and (not phone or (not phone_division and picked.get("phone_division"))):
                            phone = picked["phone"]
                            phone_division = picked.get("phone_division", "")
                        visited.add(website_url)
                    except Exception:
                        pass

                for link in result_links:
                    href = await link.get_attribute("href")
                    if not href:
                        continue
                    if "google.com" in href:
                        continue
                    if href in visited:
                        continue
                    visited.add(href)

                    try:
                        await page.goto(href, wait_until="domcontentloaded", timeout=20000)
                        await asyncio.sleep(1)
                        body_text = await page.evaluate("document.body ? document.body.innerText : ''")
                        if body_text:
                            picked = self._pick_preferred_contacts_from_text(body_text)
                            if picked.get("email") and (not email or (not email_division and picked.get("email_division"))):
                                email = picked["email"]
                                email_division = picked.get("email_division", "")
                            if picked.get("phone") and (not phone or (not phone_division and picked.get("phone_division"))):
                                phone = picked["phone"]
                                phone_division = picked.get("phone_division", "")
                        if phone and email:
                            break
                    except Exception:
                        continue
                    finally:
                        if len(visited) >= 5:
                            break
        except Exception as e:
            print(f"    ⚠️  Error searching web: {e}")
        finally:
            try:
                await page.close()
            except Exception:
                pass

        return {"phone": phone, "email": email, "phone_division": phone_division, "email_division": email_division}

    async def scrape(self, query: str) -> List[Dict]:
        """Main scraping function"""
        print("🚀 Initializing browser...")
        await self.init_browser()
        
        if not self.page:
            raise Exception("Failed to initialize browser")

        try:
            # Encode query for URL
            encoded_query = urllib.parse.quote(query)
            search_url = f"https://www.google.com/maps/search/{encoded_query}"
            
            # Navigate to Google Maps Search directly
            print(f"🗺️  Navigating to Google Maps for: '{query}'...")
            await self.page.goto(search_url, wait_until="domcontentloaded", timeout=60000)
            await self.random_delay()
            
            # Wait for results to load
            print("⏳ Waiting for results to load...")
            await self.page.wait_for_selector('[data-result-index], [role="article"], .fontHeadlineSmall', timeout=20000)
            await self.random_delay()

            # Scroll to load all results
            await self.scroll_to_load_all_results()

            # Extract business information
            print("🏪 Extracting business information...")
            results = []

            business_locator = self.page.locator('div[role="article"]')
            businesses_count = await business_locator.count()
            if businesses_count == 0:
                business_locator = self.page.locator('[data-result-index]')
                businesses_count = await business_locator.count()

            if businesses_count == 0:
                print("⚠️  No businesses found")
                return results

            print(f"📋 Found {businesses_count} businesses")

            limit = min(self.max_results, businesses_count)
            for i in range(limit):
                try:
                    print(f"🔍 Processing business {i + 1}/{businesses_count}")
                    
                    business = business_locator.nth(i)
                    for attempt in range(2):
                        try:
                            await business.scroll_into_view_if_needed()
                            await business.click()
                            break
                        except Exception as e:
                            if attempt == 0 and "not attached" in str(e).lower():
                                await asyncio.sleep(0.5)
                                continue
                            raise

                    await self.random_delay()
                    business_handle = await business.element_handle()
                    if not business_handle:
                        continue
                    
                    # Extract basic info from the list item
                    business_data = await self.extract_business_info(business_handle)
                    
                    # If name is still missing or looks like a block of text, skip it
                    if not business_data["name"] or len(business_data["name"]) > 100:
                        continue
                        
                    # Try to get additional details from the opened panel
                    try:
                        # Wait for details panel to open (use a more specific selector)
                        await self.page.wait_for_selector('div[role="main"][aria-label]', timeout=10000)
                        
                        # Add a small delay to let content render
                        await asyncio.sleep(1)
                        
                        # Extract website
                        website_selectors = [
                            'a[data-item-id="authority"]',
                            'a[aria-label*="website" i]',
                            'a[aria-label*="situs" i]'
                        ]
                        
                        for selector in website_selectors:
                            try:
                                website_element = await self.page.query_selector(selector)
                                if website_element:
                                    website_href = await website_element.get_attribute("href")
                                    if website_href and "google" not in website_href:
                                        business_data["website"] = website_href
                                        break
                            except:
                                continue

                        email_selectors = [
                            'a[href^="mailto:"]',
                            '[data-item-id^="email:"]',
                            'button[data-item-id^="email:"]',
                            'div:has-text("@")'
                        ]

                        for selector in email_selectors:
                            try:
                                email_element = await self.page.query_selector(selector)
                                if email_element:
                                    href = await email_element.get_attribute("href")
                                    if href and href.lower().startswith("mailto:"):
                                        email_val = href.split(":", 1)[1].split("?", 1)[0].strip()
                                        if email_val and re.match(r'^[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}$', email_val, re.IGNORECASE):
                                            business_data["email"] = email_val
                                            break

                                    text_val = await email_element.text_content() or await email_element.get_attribute("aria-label")
                                    if text_val and "@" in text_val:
                                        match = re.search(r'[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}', text_val, re.IGNORECASE)
                                        if match:
                                            business_data["email"] = match.group(0).strip()
                                            break
                            except:
                                continue

                        if not business_data["email"] and business_data["website"]:
                            business_data["email"] = await self._search_website_for_email(business_data["website"])

                        if business_data["email"] and not business_data["number_division"]:
                            business_data["number_division"] = self._infer_number_division(business_data["email"])
                        
                        # Extract phone
                        phone_selectors = [
                            'button[data-item-id^="phone:tel:"]',
                            'button[aria-label*="phone" i]',
                            'button[aria-label*="telepon" i]',
                            'div:has-text("+62")',
                            'div:has-text("08")',
                            'div:has-text("021")'
                        ]
                        
                        for selector in phone_selectors:
                            try:
                                phone_element = await self.page.query_selector(selector)
                                if phone_element:
                                    # First check the data-item-id which contains the clean phone number
                                    data_id = await phone_element.get_attribute("data-item-id")
                                    if data_id and data_id.startswith("phone:tel:"):
                                        business_data["phone"] = data_id.replace("phone:tel:", "").strip()
                                        break
                                        
                                    phone_text = await phone_element.text_content() or await phone_element.get_attribute("aria-label")
                                    if phone_text:
                                        # Look for phone number patterns
                                        phone_match = re.search(r'(?:\+62|0)[\d\s\-]{8,15}', phone_text)
                                        if phone_match:
                                            business_data["phone"] = phone_match.group(0).strip()
                                            break
                            except:
                                continue
                                
                    except Exception as e:
                        print(f"⚠️  Could not extract additional details: {e}")
                    
                    contact_needed = False
                    clean_phone = business_data["phone"].replace(" ", "").replace("-", "") if business_data["phone"] else ""
                    if (not business_data["phone"] or clean_phone.startswith("021") or clean_phone.startswith("+6221") or not business_data["email"] or not business_data["number_division"]):
                        contact_needed = True

                    if contact_needed:
                        found = await self._search_web_for_contacts(business_data["name"], website_url=business_data.get("website", ""))

                        if found.get("email"):
                            if not business_data["email"]:
                                business_data["email"] = found["email"]
                            elif not business_data["number_division"] and found.get("email_division"):
                                business_data["email"] = found["email"]

                        if found.get("phone"):
                            if not business_data["phone"] or clean_phone.startswith("021") or clean_phone.startswith("+6221"):
                                business_data["phone"] = found["phone"]

                        if not business_data["number_division"]:
                            business_data["number_division"] = found.get("phone_division") or found.get("email_division") or ""

                    clean_phone = business_data["phone"].replace(" ", "").replace("-", "") if business_data["phone"] else ""
                    if clean_phone.startswith("021") or clean_phone.startswith("+6221"):
                        business_data["phone"] = ""
                                
                    # Close the details panel if possible
                    try:
                        close_buttons = await self.page.query_selector_all('button[aria-label*="back"], button[aria-label*="close"]')
                        for button in close_buttons:
                            if await button.is_visible():
                                await button.click()
                                break
                    except:
                        pass
                    
                    results.append(business_data)
                    await self.random_delay()
                    
                except Exception as e:
                    print(f"⚠️  Error processing business {i + 1}: {e}")
                    continue

            return results

        except Exception as e:
            print(f"❌ Error during scraping: {e}")
            return []

    def save_to_csv(self, data: List[Dict], filename: str):
        """Save results to CSV file"""
        df = pd.DataFrame(data)
        df.insert(0, "no", range(1, len(df) + 1))
        df.columns = [str(c).upper() for c in df.columns]
        df.to_csv(filename, index=False, encoding='utf-8')

    def save_to_excel(self, data: List[Dict], filename: str):
        """Save results to Excel file"""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment

        df = pd.DataFrame(data)
        df.insert(0, "no", range(1, len(df) + 1))
        df.columns = [str(c).upper() for c in df.columns]

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Results")
            ws = writer.book["Results"]

            header_fill = PatternFill(fill_type="solid", fgColor="FFFFA500")
            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(vertical="center", wrap_text=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            body_alignment = Alignment(vertical="top", wrap_text=True)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = body_alignment

            max_width = 60
            min_width = 6
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[col_letter]:
                    value = "" if cell.value is None else str(cell.value)
                    for line in value.splitlines():
                        if len(line) > max_len:
                            max_len = len(line)
                width = max(min_width, min(max_width, max_len + 2))
                ws.column_dimensions[col_letter].width = width

    def analyze_with_ai(self, results: List[Dict]) -> str:
        """Analyze the scraped results using Groq AI model"""
        if not os.getenv("GROQ_API_KEY"):
            raise EnvironmentError(
                "GROQ_API_KEY not found in environment variables."
            )

        print("🤖 Analyzing results with AI (Groq)...")
        try:
            llm = ChatGroq(
                model="meta-llama/llama-4-scout-17b-16e-instruct",
                temperature=0
            )

            # Prepare data for AI (limit to top 15 to avoid token limits)
            sample_data = json.dumps(results[:15], indent=2)

            messages = [
                SystemMessage(content="You are an expert data analyst. Analyze the provided Google Maps scraping results. Give a brief summary of the best options based on ratings and review counts, highlight any interesting patterns, and suggest the overall top 3 places."),
                HumanMessage(content=f"Here are the top scraped results:\n{sample_data}\n\nPlease provide a structured summary and recommend the best places.")
            ]

            response = llm.invoke(messages)
            return response.content
        except Exception as e:
            print(f"❌ Error during AI analysis: {e}")
            return f"Failed to analyze data with AI: {e}"

    async def close(self):
        """Clean up browser resources"""
        if self.browser:
            await self.browser.close()
